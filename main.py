import os
import openpyxl
import win32com.client as win32

cwd = os.getcwd()

workbook = openpyxl.load_workbook(os.path.join(cwd, "Financial_Data.xlsx"))

sheet = workbook["Email_List"]

outlook = win32.Dispatch('outlook.application')

for i in range(2, sheet.max_row + 1):

    attachment = sheet.cell(row=i, column=1).value
    attachment_path = os.path.join(cwd, "Attachments", attachment)
    if not os.path.exists(attachment_path):
        print(f"Attachment {attachment} does not exist")
        continue

    recipient_name = sheet.cell(row=i, column=2).value
    recipient_email = sheet.cell(row=i, column=3).value
    cc_email = sheet.cell(row=i, column=4).value

    mail = outlook.CreateItem(0)

    mail.To = recipient_email
    mail.CC = cc_email
    mail.Subject = f"Financial Data: {attachment}"
    mail.Body = f"Dear {recipient_name},\n\nPlease find the attached financial data for {attachment}.\n\nBest regards,\nYour Name"
    mail.Attachments.Add(attachment_path)
    mail.Display()

workbook.close()
